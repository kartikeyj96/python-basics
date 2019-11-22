def calc_shipping():
    print("calc_shipping")


from pyexcel_ods import get_data  # WHEM FILE FORMAT IS .ODS ( READING THE DATA )

data = get_data("transactions.ods")
import json

print(json.dumps(data))

print("------------------------------------------------------------------------------------------")

import openpyxl as xl  # WHEM THE FILE FORMAT IS .XLSX
from openpyxl.chart import BarChart, Reference  # USE TO ADD BARCHART

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
print(cell.value)
print(sheet.max_row)  # USE TO PRINT NUMBER OF ROWS

# UPDATING SPREADSHEET (ADDING A NEW ROW)

for row in range(2, sheet.max_row + 1):  # TOTAL ROWS ARE 4 BUT IN RANGE IT PRINTS ONLY 3 THATS WHY +1
    cell = sheet.cell(row, 3)  # USE TO SELECT THE THE VALUES OF 3RD ROW
    reduced_price = cell.value * 0.5  # A NEW ROW HAVING DATA WITH THESE CHANGES
    reduced_price_cell = sheet.cell(row, 4)  # ADDING A NEW ROW
    reduced_price_cell.value = reduced_price

for row in range(2, sheet.max_row + 1):
    another_cell = sheet.cell(row, 4)
    change_price = another_cell.value * 0.7
    changed_price_cell = sheet.cell(row, 5)
    changed_price_cell.value = change_price

# MAKING AN CHART

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)  # SELECTING THE DATE FROM ROW 2 TO 4 AND COLUMS 4
chart = BarChart()  # ADDING AN BARCHART
chart.add_data(values)  # ADDING VALUES IN BARCHART
sheet.add_chart(chart, 'f2')  # ADDING CHART TO ROW E AND COLUMN 2

value = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)
chart = BarChart()
chart.add_data(value)
sheet.add_chart(chart, 'f19')

wb.save('transactions2.xlsx')  # SAVE CHANGES IN A NEW FILE



